# menu.py

"""Database-backed menu management with import/export helpers."""

from __future__ import annotations

import os
import uuid
from io import BytesIO

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from .db import SessionLocal
from .models_tenant import Category as CategoryModel, MenuItem as MenuItemModel
from .schemas import Category, CategoryIn, Item, ItemIn

router = APIRouter()


def _category_to_schema(cat: CategoryModel) -> Category:
    return Category(id=cat.id, name=cat.name)


def _item_to_schema(item: MenuItemModel) -> Item:
    return Item(
        id=item.id,
        name=item.name,
        price=item.price,
        category_id=item.category_id,
        in_stock=item.in_stock,
        show_fssai_icon=item.show_fssai_icon,
        image_url=item.image_url,
        pending_price=item.pending_price,
    )


IMAGE_DIR = os.path.join(os.path.dirname(__file__), "static", "images")
os.makedirs(IMAGE_DIR, exist_ok=True)


# Categories


@router.post("/categories", response_model=Category)
def create_category(data: CategoryIn) -> Category:
    """Create a new category and return it."""

    with SessionLocal() as session:
        category = CategoryModel(name=data.name)
        session.add(category)
        session.commit()
        session.refresh(category)
        return _category_to_schema(category)


@router.get("/categories", response_model=list[Category])
def list_categories() -> list[Category]:
    """Return all categories."""

    with SessionLocal() as session:
        categories = session.scalars(select(CategoryModel)).all()
        return [_category_to_schema(c) for c in categories]


@router.delete("/categories/{category_id}")
def delete_category(category_id: uuid.UUID) -> None:
    """Delete a category if it exists."""

    with SessionLocal() as session:
        category = session.get(CategoryModel, category_id)
        if category:
            session.delete(category)
            session.commit()


# Menu Items


@router.post("/items", response_model=Item)
def create_item(data: ItemIn) -> Item:
    """Create a menu item."""

    with SessionLocal() as session:
        item = MenuItemModel(**data.model_dump())
        session.add(item)
        session.commit()
        session.refresh(item)
        return _item_to_schema(item)


@router.get("/items", response_model=list[Item])
def list_items(include_out_of_stock: bool = False) -> list[Item]:
    """List items, optionally including out-of-stock ones."""

    with SessionLocal() as session:
        query = select(MenuItemModel)
        if not include_out_of_stock:
            query = query.where(MenuItemModel.in_stock.is_(True))
        items = session.scalars(query).all()
        return [_item_to_schema(i) for i in items]


# Import/Export


@router.get("/items/export")
def export_items() -> StreamingResponse:
    """Export current items to an Excel sheet."""
    wb = Workbook()
    ws = wb.active
    ws.append(["name", "price", "category", "in_stock", "show_fssai_icon"])
    with SessionLocal() as session:
        stmt = (
            select(MenuItemModel, CategoryModel)
            .join(CategoryModel, MenuItemModel.category_id == CategoryModel.id, isouter=True)
        )
        for item, cat in session.execute(stmt):
            ws.append(
                [
                    item.name,
                    item.price,
                    cat.name if cat else None,
                    item.in_stock,
                    item.show_fssai_icon,
                ]
            )
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    headers = {"Content-Disposition": "attachment; filename=items.xlsx"}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.ms-excel",
        headers=headers,
    )


@router.post("/items/import")
def import_items(file: UploadFile = File(...)) -> int:
    """Import items from an uploaded Excel sheet."""
    wb = load_workbook(file.file)
    ws = wb.active
    count = 0
    with SessionLocal() as session:
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, price, category_name, *_ = row
            try:
                price = int(price)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail="Invalid price")
            cid = None
            if category_name:
                cat = session.scalar(
                    select(CategoryModel).where(CategoryModel.name == category_name)
                )
                if not cat:
                    cat = CategoryModel(name=category_name)
                    session.add(cat)
                    session.flush()
                cid = cat.id
            item = MenuItemModel(name=name, price=price, category_id=cid)
            session.add(item)
            count += 1
        session.commit()
    return count


@router.get("/items/{item_id}", response_model=Item)
def get_item(item_id: uuid.UUID) -> Item:
    """Fetch a single item by ID."""
    with SessionLocal() as session:
        item = session.get(MenuItemModel, item_id)
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        return _item_to_schema(item)


@router.put("/items/{item_id}", response_model=Item)
def update_item(item_id: uuid.UUID, data: ItemIn) -> Item:
    """Update an item; price changes are staged as pending."""
    with SessionLocal() as session:
        item = session.get(MenuItemModel, item_id)
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        update = data.model_dump()
        if update["price"] != item.price:
            item.pending_price = update["price"]
            update.pop("price")
        for key, value in update.items():
            setattr(item, key, value)
        session.add(item)
        session.commit()
        session.refresh(item)
        return _item_to_schema(item)


@router.post("/items/apply-pending")
def apply_pending_prices() -> None:
    """Apply any staged price updates."""
    with SessionLocal() as session:
        items = session.scalars(
            select(MenuItemModel).where(MenuItemModel.pending_price.is_not(None))
        ).all()
        for item in items:
            item.price = item.pending_price
            item.pending_price = None
        session.commit()


@router.delete("/items/{item_id}")
def delete_item(item_id: uuid.UUID) -> None:
    """Remove an item if present."""
    with SessionLocal() as session:
        item = session.get(MenuItemModel, item_id)
        if item:
            session.delete(item)
            session.commit()


# Images


@router.post("/items/{item_id}/image", response_model=Item)
def upload_image(item_id: uuid.UUID, file: UploadFile = File(...)) -> Item:
    """Attach an image to an item by saving the uploaded file."""
    with SessionLocal() as session:
        item = session.get(MenuItemModel, item_id)
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        filename = f"{item_id}_{file.filename}"
        path = os.path.join(IMAGE_DIR, filename)
        with open(path, "wb") as f:
            f.write(file.file.read())
        item.image_url = path
        session.add(item)
        session.commit()
        session.refresh(item)
        return _item_to_schema(item)
